"""Utility module with functions related to working with Excel."""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.wave_function_collapse.utils.utils_conversion import hex_to_rgb
from src.wave_function_collapse.constants import Dimensions, RGBColor


def load_excel_worksheet(excel_file_path: Path, worksheet_index: int = 0) -> Worksheet:
    """
    Load an Excel file and return the worksheet at the given index.

    Defaults to the first worksheet, which assumes the bitmap occupies
    the first tab of the Excel file.

    Args:
        excel_file_path (Path): Path to the .xlsx file to load.
        worksheet_index (int): Index of the worksheet to return, defaults to 0.

    Returns:
        worksheet (Worksheet): The worksheet at the given index.
    """
    workbook = load_workbook(excel_file_path)
    worksheet = workbook.worksheets[worksheet_index]

    return worksheet


def detect_bitmap_dimensions(worksheet: Worksheet) -> Dimensions:
    """
    Detect the dimensions of a bitmap in an Excel worksheet by scanning for cell borders.

    Scans the first column for a bottom border to determine the number of rows,
    and the first row for a right border to determine the number of columns.
    This assumes the bitmap has an outer border drawn around it in Excel.

    Args:
        worksheet (Worksheet): The worksheet to scan for bitmap dimensions.

    Returns:
        dimensions (Dimensions): A named tuple of (rows, cols) representing the
            detected bitmap dimensions.

    Raises:
        ValueError: If no bottom or right border is detected in the worksheet,
            indicating the bitmap border is missing or malformed.
    """
    max_row = max(
        (
            row
            for row in range(1, worksheet.max_row + 1)
            if worksheet.cell(row=row, column=1).border.bottom.border_style is not None
        ),
        default=0,
    )
    max_col = max(
        (
            col
            for col in range(1, worksheet.max_column + 1)
            if worksheet.cell(row=1, column=col).border.right.border_style is not None
        ),
        default=0,
    )

    if max_row == 0 or max_col == 0:
        raise ValueError(
            f"Could not detect bitmap dimensions in worksheet '{worksheet.title}'. "
            "Ensure the bitmap has an outer border drawn around it in Excel."
        )

    dimensions = Dimensions(rows=max_row, cols=max_col)

    return dimensions


def extract_bitmap_from_worksheet(
    worksheet: Worksheet, bitmap_dimensions: Dimensions
) -> list[list[RGBColor]]:
    """
    Extract cell fill colors from the loaded worksheet as a 2D list of RGB tuples.

    Iterates over the cells within the detected bitmap dimensions and
    converts each cell's fill color from hex to RGB.

    Args:
        worksheet (Worksheet): The worksheet to extract bitmap data from.
        bitmap_dimensions (Dimensions): The row and column extent of the bitmap
            within the worksheet.

    Returns:
        bitmap (list[list[RGBColor]]): A 2D list of RGB tuples representing the bitmap.
    """
    bitmap = []

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=bitmap_dimensions.rows,
        min_col=1,
        max_col=bitmap_dimensions.cols,
    ):
        bitmap_row = []

        for cell in row:
            color_hex = cell.fill.start_color.index
            # First two characters contain transparency/alpha — strip them before converting.
            color_rgb = hex_to_rgb(color_hex=color_hex[2:])
            bitmap_row.append(color_rgb)

        bitmap.append(bitmap_row)

    return bitmap
