"""Module for reading bitmap data from .xlsx files and exporting them as .png images."""

import logging

from pathlib import Path
from PIL import Image
from openpyxl import load_workbook

from src.wave_function_collapse.config_manager import ConfigModel
from src.wave_function_collapse.constants import Size, RGBColor, BITMAPS_DIRECTORY_PATH


class BitmapUtils:
    """Handles reading bitmap data from .xlsx files and exporting them as .png images."""

    def __init__(self, config: ConfigModel, file_name: Path) -> None:
        """
        Unpack and store config values and the target bitmap file name.

        Args:
            config (ConfigModel): The validated configuration model.
            file_name (Path): Name of the bitmap .xlsx file.
        """
        self.config = config
        self.file_name = file_name
        self.bitmaps_directory_path = BITMAPS_DIRECTORY_PATH

    def _hex_to_rgb(self, color_hex: str) -> RGBColor:
        """
        Convert a 6-character hex color string to an RGB tuple.

        1. Extract the red, green, and blue components from the hex string.
        2. Convert each component from base-16 to an integer.
        3. Return the three components as a tuple.

        Args:
            color_hex (str): A 6-character hexadecimal color string, e.g. 'FF8800'.

        Returns:
            rgb_tuple (tuple[int, int, int]): A tuple of (red, green, blue) integer values.
        """
        r = int(color_hex[0:2], 16)
        g = int(color_hex[2:4], 16)
        b = int(color_hex[4:6], 16)

        rgb_tuple = (r, g, b)

        return rgb_tuple

    def _export_bitmap_as_png(
        self,
        bitmap: list[list[RGBColor]],
        file_name: Path,
        output_file_extension: str = ".png",
    ) -> None:
        """
        Export a 2D RGB bitmap as a scaled .png image.

        1. Determine the output image dimensions by scaling the bitmap size by cell_size.
        2. Create a blank RGB image filled with the configured background color.
        3. Write each bitmap color into the corresponding block of pixels.
        4. Construct the output file path and save the image.
        5. Log the output path.

        Args:
            bitmap (list[list[RGBColor]]): A 2D list of RGB tuples representing the bitmap.
            file_name (Path): The source .xlsx file name, used to derive the output file name.
            output_file_extension (str): File extension for the exported image, defaults to '.png'.
        """
        bitmap_cell_size: int = self.config.png_bitmap.cell_size
        bitmap_dimensions = Size(len(bitmap), len(bitmap[0]))
        img_width = bitmap_cell_size * bitmap_dimensions.width
        img_height = bitmap_cell_size * bitmap_dimensions.height

        image = Image.new(
            mode="RGB",
            size=(img_width, img_height),
            color=tuple(self.config.png_bitmap.background_color),
        )

        for y, bitmap_row in enumerate(bitmap):
            for x, color in enumerate(bitmap_row):
                for dx in range(bitmap_cell_size):
                    for dy in range(bitmap_cell_size):
                        image.putpixel(
                            (x * bitmap_cell_size + dx, y * bitmap_cell_size + dy),
                            color,
                        )

        image_path = self.bitmaps_directory_path / (
            file_name.stem + output_file_extension
        )
        image.save(image_path)

        logging.info(f"Bitmap exported as {image_path}")

    def read_bitmap_from_excel(self) -> list[list[tuple[int, int, int]]]:
        """
        Read cell fill colors from the first sheet of the bitmap .xlsx file.

        1. Construct the full file path from the bitmaps directory and file name.
        2. Load the first worksheet from the .xlsx file.
        3. Retrieve the expected bitmap dimensions from config.
        4. Iterate over the cells and convert each fill color from hex to RGB.
        5. Optionally export the bitmap as a .png if configured to do so.
        6. Return the bitmap as a 2D list of RGB tuples.

        Returns:
            bitmap (list[list[tuple[int, int, int]]]): A 2D list of RGB tuples representing the bitmap.
        """

        file_path_bitmap = self.bitmaps_directory_path / self.file_name
        workbook = load_workbook(file_path_bitmap)
        worksheet = workbook.worksheets[0]  # Assume bitmap is on first worksheet/tab.

        file_name_stem = self.file_name.stem
        bitmap_width: int = self.config.bitmaps[file_name_stem].dimensions[0]
        bitmap_height: int = self.config.bitmaps[file_name_stem].dimensions[1]

        bitmap = []

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=bitmap_width,
            min_col=1,
            max_col=bitmap_height,
        ):
            bitmap_row = []

            for cell in row:
                color_hex = cell.fill.start_color.index
                # First two characters contain transparency/alpha — strip them before converting.
                color_rgb = self._hex_to_rgb(color_hex=color_hex[2:])
                bitmap_row.append(color_rgb)

            bitmap.append(bitmap_row)

        if self.config.png_bitmap.export:
            self._export_bitmap_as_png(bitmap=bitmap, file_name=self.file_name)

        return bitmap

    def create_color_mapping(self, bitmap: list[list[RGBColor]]) -> dict[RGBColor, str]:
        """
        Assign a unique single character to each distinct color in the bitmap.

        1. Iterate over every color in the bitmap row by row.
        2. If the color has not been seen before, map it to the current character.
        3. Advance the character by one Unicode codepoint for the next new color.
        4. Return the completed color-to-character mapping.

        Args:
            bitmap (list[list[RGBColor]]): A 2D list of RGB tuples representing
                the bitmap.

        Returns:
            color_mapping (dict[RGBColor, str]): A dict mapping each unique RGB
                tuple to a distinct character.
        """
        color_mapping: dict[RGBColor, str] = {}
        current_char = "A"

        for bitmap_row in bitmap:
            for cell_color in bitmap_row:
                if cell_color not in color_mapping:
                    color_mapping[cell_color] = current_char
                    # chr() converts an integer codepoint back to a character, and ord() does the reverse.
                    # Together they increment the current character by one Unicode codepoint, e.g. 'A' -> 'B' -> 'C'.
                    # This would fail if the number of distinct colors exceeded the Unicode ceiling (~1.1M),
                    current_char = chr(ord(current_char) + 1)

        return color_mapping

    def apply_color_mapping(
        self,
        bitmap: list[list[RGBColor]],
        color_mapping: dict[RGBColor, str],
    ) -> list[list[str]]:
        """
        Replace each RGB tuple in the bitmap with its mapped character.

        Args:
            bitmap (list[list[RGBColor]]): A 2D list of RGB tuples representing the bitmap.
            color_mapping (dict[RGBColor, str]): A dict mapping RGB tuples to characters.

        Returns:
            mapped_bitmap (list[list[str]]): A 2D list of characters representing the color-mapped bitmap.
        """
        mapped_bitmap = [[color_mapping[color] for color in row] for row in bitmap]

        return mapped_bitmap
